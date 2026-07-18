"""
Bi-monthly Israeli VAT report generator.

Reads income and/or expense Excel files exported from iCount,
splits items into bi-monthly periods, and generates a report Excel
with 3 worksheets per period: VAT calculation, expenses, incomes.
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.personal_config import get_income_tax_advance_rate

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Israeli VAT was 17% until 2024-12-31 and 18% from 2025-01-01. Form 874 has separate
# rows for each rate, so we classify each transaction by date.
VAT_RATE_CURRENT = 0.18
VAT_RATE_PREVIOUS = 0.17
VAT_RATE_CHANGE_DATE = datetime(2025, 1, 1)

PERIOD_MONTHS = {
    1: (1, 2),
    2: (1, 2),
    3: (3, 4),
    4: (3, 4),
    5: (5, 6),
    6: (5, 6),
    7: (7, 8),
    8: (7, 8),
    9: (9, 10),
    10: (9, 10),
    11: (11, 12),
    12: (11, 12),
}

MONTH_NAMES_HE = {
    1: "ינואר",
    2: "פברואר",
    3: "מרס",
    4: "אפריל",
    5: "מאי",
    6: "יוני",
    7: "יולי",
    8: "אוגוסט",
    9: "ספטמבר",
    10: "אוקטובר",
    11: "נובמבר",
    12: "דצמבר",
}

EQUIPMENT_KEYWORDS = ["ציוד", "רכוש קבוע"]

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Arial", size=14, bold=True)
SECTION_FONT = Font(name="Arial", size=12, bold=True)
LABEL_FONT = Font(name="Arial", size=11)
VALUE_FONT = Font(name="Arial", size=11)
TOTAL_FONT = Font(name="Arial", size=11, bold=True)
RESULT_FONT = Font(name="Arial", size=13, bold=True)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
RESULT_FILL_PAY = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
RESULT_FILL_REFUND = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
TABLE_HEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

RTL_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
RTL_ALIGN_NUM = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Required columns per file kind: semantic key -> exact header text in iCount export.
# Resolved to positions at load time so column-order changes in iCount don't break the report.
COLUMN_SPECS = {
    "income": {
        "date": "תאריך",
        "before_vat": "סה\"כ לפני מע\"מ",
        "vat": "מע\"מ",
    },
    "expenses": {
        "date": "תאריך ערך",
        "expense_type": "סוג הוצאה",
        "recognized_vat": "מע\"מ מוכר",
    },
}

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def _read_headers(path: str) -> list[str]:
    """Read the first row of the active sheet as a list of stripped header strings."""
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return [str(v).strip() if v is not None else "" for v in row]
        return []
    finally:
        wb.close()


def resolve_columns(headers: list[str], kind: str) -> dict[str, int]:
    """Resolve semantic column keys to 0-based indexes by exact header match.

    Raises ValueError if any required header is missing.
    """
    spec = COLUMN_SPECS[kind]
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for key, header in spec.items():
        try:
            resolved[key] = headers.index(header)
        except ValueError:
            missing.append(f'"{header}" (for {key})')
    if missing:
        raise ValueError(
            f"Missing required {kind} columns in iCount export: " + ", ".join(missing)
        )
    return resolved


def detect_file_type(path: str) -> str | None:
    """Detect whether a file is an income or expenses export based on its headers.

    Returns the kind whose required headers are all present, or None if neither matches.
    """
    headers = _read_headers(path)
    for kind, spec in COLUMN_SPECS.items():
        if all(h in headers for h in spec.values()):
            return kind
    return None


def parse_date(value) -> datetime | None:
    """Parse a date from either a datetime object or DD/MM/YYYY string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def get_period(dt: datetime) -> tuple[int, int, int]:
    """Return (year, start_month, end_month) for the bi-monthly period."""
    start, end = PERIOD_MONTHS[dt.month]
    return (dt.year, start, end)


def period_label(year: int, start_month: int, end_month: int) -> str:
    return f"{MONTH_NAMES_HE[start_month]}-{MONTH_NAMES_HE[end_month]} {year}"


def period_sheet_prefix(year: int, start_month: int, end_month: int) -> str:
    return f"{start_month:02d}-{end_month:02d}.{year}"


def load_icount_file(path: str, kind: str) -> dict:
    """Load an iCount export preserving all raw data and sheet metadata.

    Returns a dict with:
      - headers: list of (value, number_format) for header row
      - col_widths: dict of column letter -> width
      - rows: list of dicts with 'date', 'cells' (raw values), 'formats' (number_formats)
      - cols: dict mapping semantic key (e.g. "recognized_vat") to 0-based column index
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Resolve required columns by header text
    header_strs = [
        str(c.value).strip() if c.value is not None else "" for c in ws[1]
    ]
    cols = resolve_columns(header_strs, kind)

    # Capture column widths
    col_widths = {}
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        w = ws.column_dimensions[col_letter].width
        if w:
            col_widths[col_letter] = w

    # Capture headers (value + number_format) for verbatim copy
    headers = [(cell.value, cell.number_format) for cell in ws[1]]

    # Capture data rows with their formats
    date_col = cols["date"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cells = [c.value for c in row]
        formats = [c.number_format for c in row]
        dt = parse_date(cells[date_col])
        if dt is None:
            continue
        # Replace the date cell with a parsed datetime so the verbatim sheet contains
        # comparable date values (income exports store dates as DD/MM/YYYY strings).
        cells[date_col] = dt
        rows.append({"date": dt, "cells": cells, "formats": formats})

    rows.sort(key=lambda x: x["date"])
    return {"headers": headers, "col_widths": col_widths, "rows": rows, "cols": cols}


def load_income(path: str) -> dict:
    return load_icount_file(path, kind="income")


def load_expenses(path: str) -> dict:
    return load_icount_file(path, kind="expenses")


def extract_vat_fields_income(row_cells: list, cols: dict[str, int]) -> dict:
    """Extract VAT-relevant fields from an income row."""
    return {
        "before_vat": float(row_cells[cols["before_vat"]] or 0),
        "vat": float(row_cells[cols["vat"]] or 0),
    }


def extract_vat_fields_expense(row_cells: list, cols: dict[str, int]) -> dict:
    """Extract VAT-relevant fields from an expense row."""
    expense_type = row_cells[cols["expense_type"]] or ""
    is_equipment = any(kw in expense_type for kw in EQUIPMENT_KEYWORDS)
    return {
        "recognized_vat": float(row_cells[cols["recognized_vat"]] or 0),
        "is_equipment": is_equipment,
    }


def group_rows_by_period(rows: list[dict]) -> dict[tuple, list[dict]]:
    """Group raw rows by bi-monthly period."""
    groups: dict[tuple, list[dict]] = {}
    for row in rows:
        period = get_period(row["date"])
        groups.setdefault(period, []).append(row)
    for period_rows in groups.values():
        period_rows.sort(key=lambda x: x["date"])
    return groups


# ---------------------------------------------------------------------------
# VAT calculation
# ---------------------------------------------------------------------------


def calculate_vat(
    income_rows: list[dict],
    expense_rows: list[dict],
    income_cols: dict[str, int] | None,
    expense_cols: dict[str, int] | None,
) -> dict:
    """Calculate VAT report values for a period from raw rows."""
    # Income / sales (עסקאות)
    taxable_sales = 0.0
    taxable_sales_vat = 0.0
    exempt_sales = 0.0
    if income_cols is not None:
        for row in income_rows:
            fields = extract_vat_fields_income(row["cells"], income_cols)
            if fields["vat"] > 0:
                taxable_sales += fields["before_vat"]
                taxable_sales_vat += fields["vat"]
            else:
                exempt_sales += fields["before_vat"]

    total_output_vat = taxable_sales_vat

    # Expenses / inputs (תשומות)
    equipment_vat = 0.0
    other_vat = 0.0
    if expense_cols is not None:
        for row in expense_rows:
            fields = extract_vat_fields_expense(row["cells"], expense_cols)
            if fields["is_equipment"]:
                equipment_vat += fields["recognized_vat"]
            else:
                other_vat += fields["recognized_vat"]
    total_input_vat = equipment_vat + other_vat

    result = total_output_vat - total_input_vat

    return {
        "taxable_sales": round(taxable_sales),
        "taxable_sales_vat": round(taxable_sales_vat),
        "exempt_sales": round(exempt_sales),
        "total_output_vat": round(total_output_vat),
        "equipment_vat": round(equipment_vat),
        "other_vat": round(other_vat),
        "total_input_vat": round(total_input_vat),
        "result": round(total_output_vat - total_input_vat),
    }


def calculate_advance(turnover: float, rate: float) -> dict:
    """Calculate income tax advance (מקדמות מס הכנסה) for a period."""
    advance = round(turnover * rate)
    return {"turnover": round(turnover), "rate": rate, "advance": advance}


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------


def apply_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = RTL_ALIGN
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def apply_row_border(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def write_vat_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    period_text: str,
    vat_data: dict,
    advance_rate: float | None = None,
    income_sheet_name: str | None = None,
    expense_sheet_name: str | None = None,
    income_row_count: int = 0,
    expense_row_count: int = 0,
    income_cols: dict[str, int] | None = None,
    expense_cols: dict[str, int] | None = None,
):
    """Create the VAT calculation worksheet, mirroring the layout of Israeli form 874:
    sales split into 18% / 17% / exempt-or-0%, inputs split into 18% / 17%
    (each subdivided into equipment vs. other), then total/refund and income tax advance.

    Formulas reference column letters resolved from the source's headers via *_cols,
    and classify rows by date against VAT_RATE_CHANGE_DATE.
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.rightToLeft = True

    # Column widths (A=label, B=rate, C=amount/equipment, D=vat/other)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    NUM_FMT = '#,##0'
    PCT_FMT = '0.00%'
    DATE_BOUNDARY = (
        f"DATE({VAT_RATE_CHANGE_DATE.year},{VAT_RATE_CHANGE_DATE.month},{VAT_RATE_CHANGE_DATE.day})"
    )

    # Sheet references for formulas. Column letters are derived from the source's
    # resolved column indexes (semantic key -> 0-based index).
    inc = f"'{income_sheet_name}'" if income_sheet_name else None
    exp = f"'{expense_sheet_name}'" if expense_sheet_name else None
    inc_last = income_row_count + 1  # +1 for header row
    exp_last = expense_row_count + 1
    inc_date = get_column_letter(income_cols["date"] + 1) if income_cols else None
    inc_before = get_column_letter(income_cols["before_vat"] + 1) if income_cols else None
    inc_vat = get_column_letter(income_cols["vat"] + 1) if income_cols else None
    exp_date = get_column_letter(expense_cols["date"] + 1) if expense_cols else None
    exp_type = get_column_letter(expense_cols["expense_type"] + 1) if expense_cols else None
    exp_vat = get_column_letter(expense_cols["recognized_vat"] + 1) if expense_cols else None

    inc_date_rng = f"{inc}!{inc_date}2:{inc_date}{inc_last}" if inc else None
    inc_before_rng = f"{inc}!{inc_before}2:{inc_before}{inc_last}" if inc else None
    inc_vat_rng = f"{inc}!{inc_vat}2:{inc_vat}{inc_last}" if inc else None
    exp_date_rng = f"{exp}!{exp_date}2:{exp_date}{exp_last}" if exp else None
    exp_type_rng = f"{exp}!{exp_type}2:{exp_type}{exp_last}" if exp else None
    exp_vat_rng = f"{exp}!{exp_vat}2:{exp_vat}{exp_last}" if exp else None

    def num_cell(row, col, value):
        apply_cell(ws, row, col, value, font=VALUE_FONT,
                   alignment=RTL_ALIGN_NUM, border=THIN_BORDER, number_format=NUM_FMT)

    # --- Header ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    apply_cell(ws, 1, 1, "דוח תקופתי מע\"מ", font=HEADER_FONT, fill=HEADER_FILL,
               alignment=Alignment(horizontal="center", vertical="center"))
    for c in range(1, 5):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = Font(name="Arial", size=14, bold=True, color="FFFFFF")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    apply_cell(ws, 2, 1, f"תקופת דיווח: {period_text}", font=Font(name="Arial", size=12),
               alignment=CENTER_ALIGN)

    # --- עסקאות section: rows 4-9 ---
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
    apply_cell(ws, 4, 1, "עסקאות", font=SECTION_FONT, fill=SECTION_FILL, alignment=CENTER_ALIGN)
    for c in range(1, 5):
        ws.cell(row=4, column=c).fill = SECTION_FILL
    apply_row_border(ws, 4, 1, 4)

    headers = ["", "שיעור מס", "עסקאות (ללא מע\"מ)", "מס עסקאות"]
    for c, h in enumerate(headers, 1):
        apply_cell(ws, 5, c, h, font=TOTAL_FONT, fill=TABLE_HEADER_FILL,
                   alignment=CENTER_ALIGN, border=THIN_BORDER)

    # Rows 6-7: עסקאות חייבות at current and previous rates (date-based split)
    rate_rows = [
        (6, VAT_RATE_CURRENT, f">={DATE_BOUNDARY}"),
        (7, VAT_RATE_PREVIOUS, f"<{DATE_BOUNDARY}"),
    ]
    for row, rate, date_predicate in rate_rows:
        apply_cell(ws, row, 1, "עסקאות חייבות", font=LABEL_FONT, alignment=RTL_ALIGN, border=THIN_BORDER)
        apply_cell(ws, row, 2, f"{rate:.2%}", font=VALUE_FONT,
                   alignment=CENTER_ALIGN, border=THIN_BORDER)
        if inc:
            date_filter = f"({inc_date_rng}{date_predicate})"
            num_cell(row, 3,
                f"=SUMPRODUCT(({inc_vat_rng}>0)*{date_filter}*{inc_before_rng})")
            num_cell(row, 4,
                f"=SUMPRODUCT(({inc_vat_rng}>0)*{date_filter}*{inc_vat_rng})")
        else:
            num_cell(row, 3, 0)
            num_cell(row, 4, 0)

    # Row 8: עסקאות פטורות או בשיעור אפס
    apply_cell(ws, 8, 1, "עסקאות פטורות או בשיעור אפס", font=LABEL_FONT,
               alignment=RTL_ALIGN, border=THIN_BORDER)
    apply_cell(ws, 8, 2, "", font=VALUE_FONT, alignment=CENTER_ALIGN, border=THIN_BORDER)
    if inc:
        num_cell(8, 3, f"=SUMPRODUCT(({inc_vat_rng}=0)*{inc_before_rng})")
    else:
        num_cell(8, 3, 0)
    num_cell(8, 4, 0)

    # Row 9: סה"כ מס עסקאות (sum of D6+D7)
    apply_cell(ws, 9, 1, "סה\"כ מס עסקאות:", font=TOTAL_FONT,
               alignment=RTL_ALIGN, border=THIN_BORDER)
    for c in range(2, 4):
        apply_cell(ws, 9, c, "", border=THIN_BORDER)
    apply_cell(ws, 9, 4, "=D6+D7", font=TOTAL_FONT,
               alignment=RTL_ALIGN_NUM, border=THIN_BORDER, number_format=NUM_FMT)

    # --- תשומות section: rows 11-15 ---
    ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=4)
    apply_cell(ws, 11, 1, "תשומות", font=SECTION_FONT, fill=SECTION_FILL, alignment=CENTER_ALIGN)
    for c in range(1, 5):
        ws.cell(row=11, column=c).fill = SECTION_FILL
    apply_row_border(ws, 11, 1, 4)

    headers = ["שיעור מס", "תשומות ציוד", "תשומות אחרות", "מס תשומות"]
    for c, h in enumerate(headers, 1):
        apply_cell(ws, 12, c, h, font=TOTAL_FONT, fill=TABLE_HEADER_FILL,
                   alignment=CENTER_ALIGN, border=THIN_BORDER)

    # Rows 13-14: תשומות at current and previous rates
    input_rate_rows = [
        (13, VAT_RATE_CURRENT, f">={DATE_BOUNDARY}"),
        (14, VAT_RATE_PREVIOUS, f"<{DATE_BOUNDARY}"),
    ]
    for row, rate, date_predicate in input_rate_rows:
        apply_cell(ws, row, 1, f"{rate:.2%}", font=VALUE_FONT,
                   alignment=CENTER_ALIGN, border=THIN_BORDER)
        if exp:
            date_filter = f"({exp_date_rng}{date_predicate})"
            equipment_filter = (
                f'(ISNUMBER(SEARCH("ציוד",{exp_type_rng}))'
                f'+ISNUMBER(SEARCH("רכוש קבוע",{exp_type_rng}))>0)'
            )
            num_cell(row, 2,
                f"=SUMPRODUCT({date_filter}*{equipment_filter}*{exp_vat_rng})")
            num_cell(row, 3, f"=D{row}-B{row}")
            num_cell(row, 4, f"=SUMPRODUCT({date_filter}*{exp_vat_rng})")
        else:
            num_cell(row, 2, 0)
            num_cell(row, 3, 0)
            num_cell(row, 4, 0)

    # Row 15: סה"כ מס תשומות (sum of D13+D14)
    apply_cell(ws, 15, 1, "סה\"כ מס תשומות:", font=TOTAL_FONT,
               alignment=RTL_ALIGN, border=THIN_BORDER)
    for c in range(2, 4):
        apply_cell(ws, 15, c, "", border=THIN_BORDER)
    apply_cell(ws, 15, 4, "=D13+D14", font=TOTAL_FONT,
               alignment=RTL_ALIGN_NUM, border=THIN_BORDER, number_format=NUM_FMT)

    # --- Result row: 17 ---
    result = vat_data["result"]
    if result >= 0:
        label = "סכום לתשלום:"
        fill = RESULT_FILL_PAY
    else:
        label = "סכום להחזר:"
        fill = RESULT_FILL_REFUND

    ws.merge_cells(start_row=17, start_column=1, end_row=17, end_column=3)
    apply_cell(ws, 17, 1, label, font=RESULT_FONT, fill=fill,
               alignment=Alignment(horizontal="right", vertical="center"))
    for c in range(1, 4):
        ws.cell(row=17, column=c).fill = fill
        ws.cell(row=17, column=c).border = THIN_BORDER
    apply_cell(ws, 17, 4, "=ABS(D9-D15)", font=RESULT_FONT, fill=fill,
               alignment=RTL_ALIGN_NUM, border=THIN_BORDER, number_format=NUM_FMT)

    # --- מקדמות מס הכנסה section: rows 19-21 ---
    if advance_rate is not None:
        ws.merge_cells(start_row=19, start_column=1, end_row=19, end_column=4)
        apply_cell(ws, 19, 1, "מקדמות מס הכנסה", font=SECTION_FONT,
                   fill=SECTION_FILL, alignment=CENTER_ALIGN)
        for c in range(1, 5):
            ws.cell(row=19, column=c).fill = SECTION_FILL
        apply_row_border(ws, 19, 1, 4)

        headers = ["", "מחזור עסקאות", "שיעור מקדמה", "סכום מקדמה"]
        for c, h in enumerate(headers, 1):
            apply_cell(ws, 20, c, h, font=TOTAL_FONT, fill=TABLE_HEADER_FILL,
                       alignment=CENTER_ALIGN, border=THIN_BORDER)

        apply_cell(ws, 21, 1, "", border=THIN_BORDER)
        apply_cell(ws, 21, 2, "=C6+C7+C8", font=VALUE_FONT,
                   alignment=RTL_ALIGN_NUM, border=THIN_BORDER, number_format=NUM_FMT)
        apply_cell(ws, 21, 3, advance_rate, font=VALUE_FONT,
                   alignment=CENTER_ALIGN, border=THIN_BORDER, number_format=PCT_FMT)
        apply_cell(ws, 21, 4, "=ROUND(B21*C21,0)", font=TOTAL_FONT,
                   alignment=RTL_ALIGN_NUM, border=THIN_BORDER, number_format=NUM_FMT)


def write_verbatim_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    file_meta: dict,
    period_rows: list[dict],
):
    """Copy iCount data verbatim into a new sheet, preserving columns/formats.

    Only the header row gets distinct styling; data rows keep original formats.
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.rightToLeft = True

    headers = file_meta["headers"]
    col_widths = file_meta["col_widths"]

    # Apply column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Header row — styled distinctly
    for c, (value, _fmt) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=value)
        cell.font = TOTAL_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Data rows — verbatim values and original number formats
    for r, row_data in enumerate(period_rows, 2):
        for c, (value, fmt) in enumerate(zip(row_data["cells"], row_data["formats"]), 1):
            cell = ws.cell(row=r, column=c, value=value)
            if isinstance(value, datetime):
                cell.number_format = "DD/MM/YYYY"
            elif fmt and fmt != "General":
                cell.number_format = fmt


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    load_dotenv()

    parser = argparse.ArgumentParser(
        description="Generate bi-monthly Israeli VAT report from iCount exports"
    )
    parser.add_argument("--income", "-i", help="Path to iCount income export Excel file")
    parser.add_argument("--expenses", "-e", help="Path to iCount expenses export Excel file")
    parser.add_argument("--output", "-o", default=".", help="Output directory (default: current)")
    parser.add_argument("--advance-rate", "-a", type=float,
                        help="Income tax advance rate in %% (e.g. 12 for 12%%)")

    args = parser.parse_args()

    if not args.income and not args.expenses:
        parser.error("At least one of --income or --expenses must be provided")

    # Resolve income tax advance rate (CLI → config.personal.yaml → None)
    advance_rate = None
    if args.advance_rate is not None:
        advance_rate = args.advance_rate / 100
    else:
        cfg_rate = get_income_tax_advance_rate()
        if cfg_rate is not None:
            advance_rate = cfg_rate / 100

    # Validate file types match the flags
    for flag, path, expected in [
        ("--income", args.income, "income"),
        ("--expenses", args.expenses, "expenses"),
    ]:
        if not path:
            continue
        detected = detect_file_type(path)
        if detected and detected != expected:
            print(
                f"Error: File passed as {flag} appears to be an {detected} export, not {expected}.\n"
                f"  File: {path}\n"
                f"  Hint: Did you swap the --income and --expenses arguments?"
            )
            sys.exit(1)

    # Load data
    income_data = load_income(args.income) if args.income else None
    expense_data = load_expenses(args.expenses) if args.expenses else None

    income_rows = income_data["rows"] if income_data else []
    expense_rows = expense_data["rows"] if expense_data else []
    income_cols = income_data["cols"] if income_data else None
    expense_cols = expense_data["cols"] if expense_data else None

    if not income_rows and not expense_rows:
        print("No items found in the provided files.")
        sys.exit(1)

    # Group by period
    income_by_period = group_rows_by_period(income_rows)
    expense_by_period = group_rows_by_period(expense_rows)

    # Collect all periods
    all_periods = sorted(set(income_by_period.keys()) | set(expense_by_period.keys()))

    print(f"Found {len(all_periods)} reporting period(s):")
    for period in all_periods:
        label = period_label(*period)
        n_inc = len(income_by_period.get(period, []))
        n_exp = len(expense_by_period.get(period, []))
        print(f"  {label}: {n_inc} income, {n_exp} expense items")

    # Generate output
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for period in all_periods:
        prefix = period_sheet_prefix(*period)
        label = period_label(*period)

        period_income = income_by_period.get(period, [])
        period_expenses = expense_by_period.get(period, [])

        vat_data = calculate_vat(period_income, period_expenses, income_cols, expense_cols)

        expense_sheet_name = f"{prefix} הוצאות" if expense_data else None
        income_sheet_name = f"{prefix} הכנסות" if income_data else None

        write_vat_sheet(
            wb, f"{prefix} מע\"מ", label, vat_data,
            advance_rate=advance_rate,
            income_sheet_name=income_sheet_name,
            expense_sheet_name=expense_sheet_name,
            income_row_count=len(period_income),
            expense_row_count=len(period_expenses),
            income_cols=income_cols,
            expense_cols=expense_cols,
        )
        if expense_data:
            write_verbatim_sheet(wb, expense_sheet_name, expense_data, period_expenses)
        if income_data:
            write_verbatim_sheet(wb, income_sheet_name, income_data, period_income)

    # Save
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Use first and last period for filename
    first = all_periods[0]
    last = all_periods[-1]
    filename = f"vat_report_{first[0]}{first[1]:02d}-{last[0]}{last[2]:02d}.xlsx"
    output_path = output_dir / filename

    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")


if __name__ == "__main__":
    main()
