#!/usr/bin/env python3
"""
Receipt Data Consolidator - Stage 2
Processes reviewed Excel files and generates iCount-ready Excel import files
"""

import argparse
import logging
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv
import sys
import openpyxl
import xlwt
import shutil
import os
import re

# Add shared modules to path
sys.path.append(str(Path(__file__).parent))

from shared.logger import ReceiptLogger
from shared.excel_config import get_excel_config

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ReceiptConsolidator:
    """Consolidates reviewed Excel files into iCount import format"""
    
    def __init__(self, output_dir: Path, receipts_source_dir: Optional[Path] = None):
        """Initialize consolidator

        Args:
            output_dir: Directory for consolidated output files
            receipts_source_dir: Optional directory to search for original receipt files
        """
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Store receipts source directory
        self.receipts_source_dir = receipts_source_dir

        # Setup logging
        self.logs_dir = self.output_dir / 'consolidation_logs'
        self.logger = ReceiptLogger(self.logs_dir)

        # Load Excel configuration
        self.config = get_excel_config()
        
    def process_excel_files(self, excel_files: List[Path]) -> Dict[str, Any]:
        """Process multiple Excel files and generate consolidated output"""
        start_time = datetime.now()
        
        logger.info(f"Processing {len(excel_files)} Excel files for consolidation")
        
        all_receipts = []
        processing_errors = []
        
        # Process each Excel file
        for excel_file in excel_files:
            try:
                receipts = self._extract_receipts_from_excel(excel_file)
                all_receipts.extend(receipts)
                logger.info(f"Extracted {len(receipts)} receipts from {excel_file.name}")
                
            except Exception as e:
                logger.error(f"Error processing {excel_file}: {e}")
                processing_errors.append({
                    'file': str(excel_file),
                    'error': str(e)
                })
                
        if not all_receipts:
            logger.warning("No receipts extracted from Excel files")
            return {
                'status': 'error',
                'message': 'No receipts found in Excel files',
                'errors': processing_errors
            }
            
        # Generate iCount Excel
        excel_file = self._generate_icount_excel(all_receipts)

        # Copy receipt files to organized structure
        copy_stats = self._copy_receipt_files(all_receipts)

        # Generate summary
        end_time = datetime.now()
        summary = self._generate_summary(all_receipts, excel_file, processing_errors, start_time, end_time)

        # Add receipt copy statistics to summary
        summary['receipt_files'] = copy_stats

        # Log consolidation stats
        self.logger.log_processing_stats(summary)

        return summary
        
    def _extract_receipts_from_excel(self, excel_file: Path) -> List[Dict[str, Any]]:
        """Extract receipt data from a single Excel file"""
        receipts = []

        # Read using openpyxl with data_only=True to get calculated formula values
        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Convert worksheet to DataFrame
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    data.append(row)

                df = pd.DataFrame(data)

                try:
                    receipt = self._parse_worksheet(df, worksheet, sheet_name, excel_file)
                    if receipt:
                        receipts.append(receipt)
                except Exception as e:
                    logger.error(f"Error parsing worksheet {sheet_name} in {excel_file}: {e}")

        except Exception as e:
            logger.error(f"Error reading Excel file {excel_file}: {e}")

        return receipts
        
    def _parse_worksheet(self, df: pd.DataFrame, worksheet: Any, sheet_name: str, excel_file: Path) -> Optional[Dict[str, Any]]:
        """Parse a single worksheet into receipt data"""
        try:
            # Extract header information (rows 1-10)
            receipt_data = {}

            # Get field mappings from configuration
            header_mapping = self.config.get_field_mappings()

            # Extract header fields using configuration
            max_header_rows = self.config.header_max_rows
            for idx, row in df.iloc[:max_header_rows].iterrows():
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                    field_name = str(row.iloc[0]).strip()
                    field_value = row.iloc[1]

                    if field_name in header_mapping:
                        mapped_field = header_mapping[field_name]
                        receipt_data[mapped_field] = field_value

                        # For original_file field, try to extract full path from hyperlink
                        if mapped_field == 'original_file':
                            try:
                                # Cell is in row idx+1 (1-indexed), column B (2)
                                cell = worksheet.cell(row=idx+1, column=2)
                                if cell.hyperlink and cell.hyperlink.target:
                                    # Extract path from hyperlink (remove file:// prefix if present)
                                    hyperlink_target = cell.hyperlink.target
                                    if hyperlink_target.startswith('file://'):
                                        full_path = hyperlink_target[7:]  # Remove 'file://'
                                    else:
                                        full_path = hyperlink_target  # Use as-is
                                    receipt_data['original_file_full_path'] = full_path
                                    logger.debug(f"Extracted full path from hyperlink: {full_path}")
                            except Exception as e:
                                logger.debug(f"Could not extract hyperlink from cell: {e}")
                        
            # Validate required fields
            required_fields = ['number', 'vendor', 'date', 'total_incl_vat', 'category']
            missing_fields = [f for f in required_fields if not receipt_data.get(f)]
            
            if missing_fields:
                logger.warning(f"Missing required fields in {sheet_name}: {missing_fields}")
                return None
                
            # Extract line items using configuration
            line_items = []
            line_item_start = self.config.line_items_start_row
            
            if len(df) > line_item_start:
                for idx, row in df.iloc[line_item_start:].iterrows():
                    # Check if row has data (description not empty)
                    if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip():
                        line_item = {
                            'description': str(row.iloc[0]).strip(),
                            'amount_excl_vat': self._safe_float(row.iloc[1] if len(row) > 1 else 0),
                            'vat': self._safe_float(row.iloc[2] if len(row) > 2 else 0),
                            'total': self._safe_float(row.iloc[4] if len(row) > 4 else 0),
                            'deductible': self._safe_bool(row.iloc[5] if len(row) > 5 else True)
                        }
                        line_items.append(line_item)
                        
            # Add metadata
            receipt_data['line_items'] = line_items
            receipt_data['source_file'] = str(excel_file)
            receipt_data['worksheet'] = sheet_name
            receipt_data['processing_date'] = datetime.now().isoformat()
            
            return receipt_data
            
        except Exception as e:
            logger.error(f"Error parsing worksheet {sheet_name}: {e}")
            return None
            
    def _generate_icount_excel(self, receipts: List[Dict[str, Any]]) -> Path:
        """Generate Excel file in iCount import format"""

        # Prepare data for Excel
        data_rows = []

        for receipt in receipts:
            # Create one row per receipt (not per line item)
            data_rows.append(self._create_icount_row(receipt, None))

        # Clean string data to remove double quotes, handle NA values, and fix numeric formatting
        cleaned_data_rows = []
        for row in data_rows:
            cleaned_row = {}
            for key, value in row.items():
                if pd.isna(value):
                    cleaned_row[key] = ''
                elif isinstance(value, str):
                    # Remove double quotes from strings
                    cleaned_value = value.replace('"', '')

                    # Convert "NA" values in vendor ID column to empty string
                    if key == 'תז/חפ הספק' and cleaned_value.upper() == 'NA':
                        cleaned_row[key] = ''
                    else:
                        # Remove trailing .0 from numeric strings
                        if cleaned_value.endswith('.0'):
                            cleaned_value = cleaned_value[:-2]
                        cleaned_row[key] = cleaned_value
                else:
                    # Convert all non-string values to string and clean
                    value_str = str(value)
                    if value_str.endswith('.0'):
                        value_str = value_str[:-2]
                    cleaned_row[key] = value_str
            cleaned_data_rows.append(cleaned_row)

        # Create DataFrame
        df = pd.DataFrame(cleaned_data_rows)

        # Group by vendor and sort by date
        if not df.empty:
            # Sort by vendor id and name first, then by date
            df = df.sort_values(by=['תז/חפ הספק', 'שם הספק', 'תאריך האסמכתא'], na_position='last')

            logger.info(f"Sorted {len(df)} rows by vendor name and date")

        # Generate Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = self.output_dir / f'icount_import_{timestamp}.xls'

        # Save as true XLS format using xlwt
        self._save_to_xls(df, excel_file)

        logger.info(f"Generated iCount Excel: {excel_file}")
        return excel_file

    def _format_xls_output(self, worksheet, df: pd.DataFrame, workbook):
        """Format the XLS worksheet for better readability"""

        # Create formats for xlsxwriter
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter'
        })

        currency_format = workbook.add_format({'num_format': '#,##0.00'})
        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})

        # Format header row
        for col_num, header in enumerate(df.columns):
            worksheet.write(0, col_num, header, header_format)

        # Auto-adjust column widths for Hebrew text and data
        for col_num, column_name in enumerate(df.columns):
            # Calculate max length, accounting for Hebrew characters
            max_length = len(str(column_name))

            # Check data in this column
            for value in df.iloc[:, col_num]:
                if pd.notna(value):
                    length = len(str(value))
                    # Hebrew characters are typically wider, so add some padding
                    if any('\u0590' <= char <= '\u05FF' for char in str(value)):
                        length = int(length * 1.2)  # Extra width for Hebrew
                    max_length = max(max_length, length)

            # Set minimum and maximum column width
            adjusted_width = min(max(max_length + 2, 8), 50)
            worksheet.set_column(col_num, col_num, adjusted_width)

        # Set right-to-left reading order for Hebrew
        worksheet.right_to_left()

        # Format specific columns with data
        for row_num in range(len(df)):
            for col_num, header in enumerate(df.columns):
                value = df.iloc[row_num, col_num]

                # Format amount columns (סכום) as currency
                if header == 'סכום' and pd.notna(value) and isinstance(value, (int, float)):
                    worksheet.write(row_num + 1, col_num, value, currency_format)

                # Format date columns
                elif any(date_word in header for date_word in ['תאריך', 'שולמה']) and pd.notna(value):
                    worksheet.write(row_num + 1, col_num, value, date_format)

        # Freeze the header row
        worksheet.freeze_panes(1, 0)

    def _save_to_xls(self, df: pd.DataFrame, excel_file: Path):
        """Save DataFrame to true XLS format using xlwt"""
        # Create workbook and worksheet
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet('iCount Import')

        # Create styles for formatting
        header_style = xlwt.XFStyle()
        header_font = xlwt.Font()
        header_font.bold = True
        header_font.colour_index = xlwt.Style.colour_map['white']
        header_pattern = xlwt.Pattern()
        header_pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        header_pattern.pattern_fore_colour = xlwt.Style.colour_map['blue']
        header_style.font = header_font
        header_style.pattern = header_pattern
        header_style.alignment.horz = xlwt.Alignment.HORZ_CENTER

        # Write header row with formatting
        for col_idx, col_name in enumerate(df.columns):
            worksheet.write(0, col_idx, str(col_name), header_style)

        # Write data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 1):
            for col_idx, value in enumerate(row):
                if pd.isna(value) or str(value).lower() == 'nan':
                    worksheet.write(row_idx, col_idx, '')
                else:
                    value_str = str(value)
                    # Remove .0 from end if present
                    if value_str.endswith('.0'):
                        value_str = value_str[:-2]
                    worksheet.write(row_idx, col_idx, value_str)

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df.columns):
            # Calculate max length
            max_length = len(str(col_name))
            for value in df.iloc[:, col_idx]:
                if pd.notna(value):
                    length = len(str(value))
                    # Hebrew characters are typically wider
                    if any('\u0590' <= char <= '\u05FF' for char in str(value)):
                        length = int(length * 1.2)
                    max_length = max(max_length, length)

            # Set column width (xlwt uses 256 units per character)
            adjusted_width = min(max(max_length + 2, 8), 50) * 256
            worksheet.col(col_idx).width = adjusted_width

        # Save workbook
        workbook.save(str(excel_file))

    def _sanitize_filename(self, text: str) -> str:
        """Sanitize text for use in filename"""
        if not text:
            return "unknown"

        # Remove or replace invalid filename characters
        text = re.sub(r'[<>:"/\\|?*]', '_', text)
        # Remove extra whitespace and Hebrew characters that might cause issues
        text = re.sub(r'\s+', '_', text.strip())
        # Remove quotes and other problematic characters
        text = re.sub(r'["\'\[\](){}]', '', text)
        # Limit length
        text = text[:50] if text else "unknown"

        return text

    def _find_receipt_file(self, original_filename: str, full_path: Optional[str] = None) -> Optional[Path]:
        """Find receipt file using simplified search logic

        Search order:
        1. Try full path from Excel hyperlink (if available)
        2. Try current working directory
        3. Try current working directory / receipts
        4. Try user-specified receipts source directory (if provided)
        """
        if not original_filename:
            return None

        # Extract just the filename
        filename_only = Path(original_filename).name
        current_dir = Path.cwd()

        # 1. Try full path from hyperlink first
        if full_path:
            full_path_obj = Path(full_path)
            if full_path_obj.exists() and full_path_obj.is_file():
                logger.debug(f"Found receipt file at full path: {full_path}")
                return full_path_obj

        # 2. Try current working directory
        current_file = current_dir / filename_only
        if current_file.exists() and current_file.is_file():
            logger.debug(f"Found receipt file in current directory: {current_file}")
            return current_file

        # 3. Try current working directory / receipts
        receipts_file = current_dir / "receipts" / filename_only
        if receipts_file.exists() and receipts_file.is_file():
            logger.debug(f"Found receipt file in ./receipts: {receipts_file}")
            return receipts_file

        # 4. Try user-specified receipts source directory
        if self.receipts_source_dir:
            source_file = self.receipts_source_dir / filename_only
            if source_file.exists() and source_file.is_file():
                logger.debug(f"Found receipt file in source directory: {source_file}")
                return source_file

        return None

    def _copy_receipt_files(self, receipts: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Copy receipt files to organized folder structure"""
        # Create receipts folder
        receipts_folder = self.output_dir / "receipts"
        receipts_folder.mkdir(exist_ok=True)

        copy_stats = {
            "total_receipts": len(receipts),
            "files_found": 0,
            "files_copied": 0,
            "files_missing": 0,
            "copy_errors": []
        }

        for i, receipt in enumerate(receipts, 1):
            try:
                # Extract receipt information
                original_filename = receipt.get('original_file', '')
                receipt_date = receipt.get('date', '')
                vendor_name = receipt.get('vendor', '')
                receipt_id = receipt.get('number', f"{i:03d}")  # Use actual receipt number from column H

                # Format date
                date_str = ''
                if receipt_date:
                    try:
                        if isinstance(receipt_date, str):
                            # Try different date formats
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d.%m.%Y']:
                                try:
                                    dt = datetime.strptime(receipt_date, fmt)
                                    date_str = dt.strftime('%Y%m%d')
                                    break
                                except ValueError:
                                    continue
                        elif hasattr(receipt_date, 'strftime'):
                            date_str = receipt_date.strftime('%Y%m%d')
                    except Exception as e:
                        logger.warning(f"Error formatting date {receipt_date}: {e}")

                if not date_str:
                    date_str = datetime.now().strftime('%Y%m%d')

                # Find source file
                if original_filename:
                    # Try to get full path from hyperlink if available
                    full_path = receipt.get('original_file_full_path')
                    source_file = self._find_receipt_file(original_filename, full_path)
                    if source_file:
                        copy_stats["files_found"] += 1

                        # Generate new filename
                        sanitized_vendor = self._sanitize_filename(vendor_name)
                        sanitized_receipt_id = self._sanitize_filename(str(receipt_id))
                        file_extension = source_file.suffix
                        new_filename = f"{date_str}_{sanitized_receipt_id}__{sanitized_vendor}{file_extension}"

                        # Handle duplicate names
                        target_file = receipts_folder / new_filename
                        counter = 1
                        while target_file.exists():
                            name_part = new_filename.rsplit('.', 1)[0]
                            ext_part = f".{new_filename.rsplit('.', 1)[1]}" if '.' in new_filename else ''
                            new_filename = f"{name_part}_{counter}{ext_part}"
                            target_file = receipts_folder / new_filename
                            counter += 1

                        # Copy file
                        shutil.copy2(source_file, target_file)
                        copy_stats["files_copied"] += 1
                        logger.info(f"Copied receipt file: {original_filename} -> {new_filename}")

                    else:
                        copy_stats["files_missing"] += 1
                        logger.warning(f"Receipt file not found: {original_filename} (Receipt ID: {receipt_id})")
                else:
                    copy_stats["files_missing"] += 1
                    logger.warning(f"No original filename specified for receipt {receipt_id}")

            except Exception as e:
                copy_stats["copy_errors"].append({
                    "receipt": receipt_id,
                    "filename": receipt.get('original_file', ''),
                    "error": str(e)
                })
                logger.error(f"Error copying receipt file for receipt {receipt_id}: {e}")

        logger.info(f"Receipt file copying complete: {copy_stats['files_copied']}/{copy_stats['total_receipts']} files copied")
        return copy_stats

    def _create_icount_row(self, receipt: Dict[str, Any], line_item: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """Create a single iCount row for import"""

        if line_item:
            amount = line_item['total']
        else:
            amount = receipt.get('total_incl_vat', 0)

        # Get category and append VAT 0 indicator if applicable
        category = receipt.get('category', '')
        vat_amount = receipt.get('vat_amount', 0)
        if vat_amount == 0 and category:
            category = f"{category} (ללא מע״מ)"

        # Format date in yyyy-mm-dd format
        date_str = self._format_date_icount(receipt.get('date', ''))

        # Map currency to numeric code
        currency_code = self._map_currency_to_code(receipt.get('currency', ''))

        # Map document type to iCount format
        doc_type = self._map_document_type_to_icount(receipt.get('document_type', ''))

        # Return all 15 columns required by iCount
        return {
            'תז/חפ הספק': receipt.get('vendor_id', ''),  # Column A
            'שם הספק': receipt.get('vendor', ''),  # Column B
            'שם סוג הוצאה': category,  # Column C
            'סכום': amount,  # Column D
            'מטבע': currency_code,  # Column E
            'שער': '',  # Column F
            'סוג מסמך': doc_type,  # Column G
            'מספר מסמך': receipt.get('number', ''),  # Column H
            'תאריך האסמכתא': date_str,  # Column I
            'תאריך התשלום': date_str,  # Column J
            'ההוצאה שולמה': '1',  # Column K
            'שולמה בתאריך': date_str,  # Column L
            'תאריך דיווח שונה': '',  # Column M
            'לקוח': '',  # Column N
            'פרויקט': ''  # Column O
        }
        
    def _format_date(self, date_value: Any) -> str:
        """Format date for iCount import"""
        if pd.isna(date_value):
            return ''

        try:
            # Try to parse as datetime
            if isinstance(date_value, str):
                # Handle different date formats
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d.%m.%Y']:
                    try:
                        dt = datetime.strptime(date_value, fmt)
                        return dt.strftime('%d/%m/%Y')
                    except ValueError:
                        continue

            elif hasattr(date_value, 'strftime'):
                return date_value.strftime('%d/%m/%Y')

        except Exception as e:
            logger.warning(f"Error formatting date {date_value}: {e}")

        return str(date_value)

    def _format_date_icount(self, date_value: Any) -> str:
        """Format date for iCount import in yyyy-mm-dd format"""
        if pd.isna(date_value):
            return ''

        try:
            # Try to parse as datetime
            if isinstance(date_value, str):
                # Handle different date formats
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d.%m.%Y']:
                    try:
                        dt = datetime.strptime(date_value, fmt)
                        return dt.strftime('%Y-%m-%d')
                    except ValueError:
                        continue

            elif hasattr(date_value, 'strftime'):
                return date_value.strftime('%Y-%m-%d')

        except Exception as e:
            logger.warning(f"Error formatting date {date_value}: {e}")

        return str(date_value)

    def _map_currency_to_code(self, currency: str) -> str:
        """Map currency name to iCount numeric code"""
        # Currency mapping: 1=EUR, 2=USD, 3=JPY, 4=GBP, 5=ILS
        # Empty means ILS (default)
        currency_mapping = {
            'ILS': '',  # Default, no need to specify
            'שקל': '',
            '₪': '',
            'NIS': '',
            'USD': '2',
            'דולר': '2',
            '$': '2',
            'EUR': '1',
            'אירו': '1',
            '€': '1',
            'GBP': '4',
            'לירה סטרלינג': '4',
            '£': '4',
            'JPY': '3',
            'ין': '3',
            '¥': '3'
        }

        if not currency:
            return ''  # Default to ILS

        # Try to find currency in mapping
        currency_upper = str(currency).upper().strip()
        return currency_mapping.get(currency_upper, currency_mapping.get(currency, ''))

    def _map_document_type_to_icount(self, doc_type: str) -> str:
        """Map document type to iCount format"""
        # iCount expects: 'invrec', 'receipt', 'invoice', or 'deal'
        doc_type_mapping = {
            'חשבונית': 'invoice',
            'קבלה': 'receipt',
            'חשבונית+קבלה': 'invrec',
            'חשבונית מס קבלה': 'invrec',
            'invoice': 'invoice',
            'receipt': 'receipt',
            'invoice+receipt': 'invrec',
            'deal': 'deal'
        }

        if not doc_type:
            return 'invrec'  # Default

        doc_type_lower = str(doc_type).lower().strip()
        return doc_type_mapping.get(doc_type_lower, doc_type_mapping.get(doc_type, 'invrec'))
        
    def _safe_float(self, value: Any) -> float:
        """Safely convert value to float"""
        if pd.isna(value):
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
            
    def _safe_bool(self, value: Any) -> bool:
        """Safely convert value to boolean"""
        if pd.isna(value):
            return True
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['true', '1', 'yes', 'כן']
        try:
            return bool(int(value))
        except (ValueError, TypeError):
            return True
            
    def _generate_summary(
        self,
        receipts: List[Dict[str, Any]],
        excel_file: Path,
        errors: List[Dict[str, str]],
        start_time: datetime,
        end_time: datetime
    ) -> Dict[str, Any]:
        """Generate processing summary"""
        
        total_amount = sum(receipt.get('total_incl_vat', 0) for receipt in receipts)
        
        summary = {
            'timestamp': datetime.now().isoformat(),
            'total_receipts': len(receipts),
            'total_amount': total_amount,
            'excel_file': str(excel_file),
            'processing_time_seconds': (end_time - start_time).total_seconds(),
            'errors': errors
        }
        
        # Category breakdown
        categories = {}
        for receipt in receipts:
            category = receipt.get('category', 'Unknown')
            categories[category] = categories.get(category, 0) + 1
            
        summary['category_breakdown'] = categories
        
        return summary


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Consolidate reviewed Excel files into iCount import format'
    )
    parser.add_argument(
        'excel_files',
        nargs='+',
        type=Path,
        help='Excel files to consolidate'
    )
    parser.add_argument(
        '--output',
        type=Path,
        default=Path('./receipts_consolidated'),
        help='Output directory (default: ./receipts_consolidated)'
    )
    parser.add_argument(
        '--receipts-source-dir',
        type=Path,
        default=None,
        help='Directory to search for original receipt files (optional)'
    )

    args = parser.parse_args()
    
    # Validate input files
    valid_files = []
    for excel_file in args.excel_files:
        if excel_file.exists() and excel_file.suffix.lower() in ['.xlsx', '.xls']:
            valid_files.append(excel_file)
        else:
            logger.warning(f"Skipping invalid file: {excel_file}")
            
    if not valid_files:
        logger.error("No valid Excel files found")
        sys.exit(1)
        
    # Create timestamp-based output directory
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = args.output / f'consolidation_{timestamp}'

    # Initialize consolidator
    consolidator = ReceiptConsolidator(output_dir, receipts_source_dir=args.receipts_source_dir)
    
    # Process files
    logger.info(f"Starting consolidation of {len(valid_files)} Excel files")
    summary = consolidator.process_excel_files(valid_files)
    
    # Print summary
    print("\n" + "="*60)
    print("CONSOLIDATION COMPLETE")
    print("="*60)
    print(f"Total receipts: {summary.get('total_receipts', 0)}")
    print(f"Total amount: ₪{summary.get('total_amount', 0):,.2f}")
    print(f"Processing time: {summary.get('processing_time_seconds', 0):.1f} seconds")
    
    if summary.get('excel_file'):
        print(f"\nGenerated Excel file: {summary['excel_file']}")
        
    if summary.get('category_breakdown'):
        print("\nCategory breakdown:")
        for category, count in summary['category_breakdown'].items():
            print(f"  {category}: {count} receipts")
            
    if summary.get('errors'):
        print(f"\nErrors encountered: {len(summary['errors'])}")
        for error in summary['errors']:
            print(f"  - {error['file']}: {error['error']}")

    # Print receipt file copying statistics
    if summary.get('receipt_files'):
        receipt_stats = summary['receipt_files']
        print(f"\nReceipt files:")
        print(f"  Files copied: {receipt_stats['files_copied']}/{receipt_stats['total_receipts']}")
        print(f"  Files found: {receipt_stats['files_found']}")
        print(f"  Files missing: {receipt_stats['files_missing']}")

        if receipt_stats['copy_errors']:
            print(f"  Copy errors: {len(receipt_stats['copy_errors'])}")

    print(f"\nOutput directory: {output_dir}")
    if summary.get('receipt_files', {}).get('files_copied', 0) > 0:
        print(f"Receipt files directory: {output_dir / 'receipts'}")
    
    # Save summary to JSON
    summary_path = output_dir / f'consolidation_summary_{timestamp}.json'
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
        
    print(f"Summary saved to: {summary_path}")


if __name__ == '__main__':
    main()