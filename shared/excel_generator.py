# Copyright (c) 2025 Michael Litvin
# Licensed under AGPL-3.0-or-later - see LICENSE file for details
"""Excel file generation for receipt processing"""

import logging
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.workbook.defined_name import DefinedName

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel files for receipt data"""

    def __init__(self, categories_file_path: Path):
        """Initialize with categories file path"""
        self.categories_file_path = categories_file_path
        # Import here to avoid circular imports
        from .excel_config import get_excel_config
        self.config = get_excel_config()
        # Load categories after config is set
        self.categories = self._load_categories()
        
    def create_batch_workbook(self, receipts: List[Dict[str, Any]], images_dir: Path) -> Workbook:
        """Create workbook with multiple receipt worksheets"""
        wb = Workbook()

        # Remove default sheet
        default_sheet = self.config.get_default_sheet_name()
        if default_sheet in wb.sheetnames:
            wb.remove(wb[default_sheet])

        # Create hidden sheet for categories (to avoid 255 char limit in data validation)
        if self.categories:
            self._create_categories_sheet(wb)

        # Create worksheet for each receipt
        for idx, receipt in enumerate(receipts, 1):
            ws_name = self.config.get_worksheet_name(idx)
            ws = wb.create_sheet(title=ws_name)
            self._create_receipt_worksheet(ws, receipt, images_dir)

        return wb

    def _create_categories_sheet(self, wb: Workbook):
        """Create a hidden sheet with categories and define a named range"""
        ws = wb.create_sheet(title="_Categories")

        # Write categories to column A
        for idx, category in enumerate(self.categories, 1):
            ws.cell(row=idx, column=1, value=category)

        # Hide the sheet
        ws.sheet_state = 'hidden'

        # Create named range for categories
        # Format: '_Categories'!$A$1:$A$N where N is the number of categories
        last_row = len(self.categories)
        ref = f"'_Categories'!$A$1:$A${last_row}"
        defn = DefinedName("CategoryList", attr_text=ref)
        wb.defined_names.add(defn)
        
    def _create_receipt_worksheet(self, ws: Worksheet, receipt: Dict[str, Any], images_dir: Path):
        """Create a single receipt worksheet with data and image"""

        # Set column widths from configuration
        ws.column_dimensions['A'].width = self.config.get_header_column_width('field_name')
        ws.column_dimensions['B'].width = self.config.get_header_column_width('value')
        ws.column_dimensions['C'].width = self.config.get_header_column_width('verification')
        ws.column_dimensions['D'].width = self.config.get_header_column_width('notes')

        # For line items - use default width from config
        line_item_width = self.config.config['column_widths']['line_items_section']['default']
        for col in range(5, 8):  # E, F, G
            ws.column_dimensions[get_column_letter(col)].width = line_item_width

        # Image columns (H onwards)
        ws.column_dimensions['H'].width = self.config.config['column_widths']['image_section']['image_column']
        
        # Add header section
        self._add_header_section(ws, receipt)
        
        # Add line items section
        self._add_line_items_section(ws, receipt.get('line_items', []))
        
        # Add image if available
        self._add_receipt_image(ws, receipt, images_dir)
        
        # Add validation and formatting
        self._add_validation_and_formatting(ws, receipt)
        
    def _add_header_section(self, ws: Worksheet, receipt: Dict[str, Any]):
        """Add header information section"""
        # Add title row using dynamic config
        title_cells = self.config.get_title_cells()
        titles = self.config.get_header_titles()

        for cell_ref, title in zip(title_cells, titles):
            ws[cell_ref] = title
            ws[cell_ref].font = Font(bold=True)
            ws[cell_ref].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Add receipt info
        receipt_info = receipt.get('receipt_info', {})
        amounts = receipt.get('amounts', {})
        classification = receipt.get('classification', {})

        row = self.config.header_start_row
        for hebrew_name, field_key in self.config.get_header_fields():
            field_cell = self.config.get_cell_reference(row, self.config.header_field_column)
            value_cell = self.config.get_cell_reference(row, self.config.header_value_column)
            ws[field_cell] = hebrew_name

            # Get value based on field
            if field_key in receipt_info:
                value = receipt_info[field_key]
            elif field_key in amounts:
                value = amounts[field_key]
            elif field_key == 'category':
                value = classification.get('category', '')
            else:
                value = ''

            # Special handling for document type
            if field_key == 'document_type':
                ws[value_cell] = self.config.get_document_type_mapping(value)
                # Add dropdown validation
                document_types = self.config.get_document_types()
                dv = DataValidation(type="list", formula1='"' + ','.join(document_types) + '"')
                dv.add(ws[value_cell])
                ws.add_data_validation(dv)
            elif field_key == 'category':
                ws[value_cell] = value
                # Add category dropdown using named range (avoids 255 char limit)
                if self.categories:
                    dv = DataValidation(type="list", formula1="=CategoryList")
                    dv.add(ws[value_cell])
                    ws.add_data_validation(dv)
            elif field_key == 'original_file':
                # Add hyperlink to original file with filename as display text
                original_file_path = receipt_info.get('original_file', '')
                if original_file_path:
                    filename = Path(original_file_path).name
                    ws[value_cell] = filename
                    # Create proper file:// URL for absolute path
                    # Excel requires file:///C:/path format (three slashes, forward slashes)
                    resolved_path = Path(original_file_path).resolve().as_posix()
                    file_url = f"file:///{resolved_path}"
                    ws[value_cell].hyperlink = file_url
                    ws[value_cell].font = Font(color=self.config.get_color('hyperlink'), underline="single")
                else:
                    ws[value_cell] = ''
            elif field_key == 'reasoning':
                # Make reasoning cell multiline with text wrapping
                ws[value_cell] = value
                ws[value_cell].alignment = Alignment(wrap_text=True, vertical='top')
                # Make the row taller for reasoning
                ws.row_dimensions[row].height = self.config.config['formatting']['line_items_section']['reasoning_cell_height']
            else:
                ws[value_cell] = value

            # Add verification formula for total_incl_vat to check if it equals excl + vat
            if field_key == 'total_incl_vat':
                verify_cell = self.config.get_cell_reference(row, self.config.header_value_column + 1)
                excl_cell = self.config.get_cell_reference(row-2, self.config.header_value_column)
                vat_cell = self.config.get_cell_reference(row-1, self.config.header_value_column)
                ws[verify_cell] = f'={excl_cell}+{vat_cell}'

            row += 1
            
    def _add_line_items_section(self, ws: Worksheet, line_items: List[Dict[str, Any]]):
        """Add line items table"""
        start_row = self.config.line_items_header_row
        headers = self.config.get_line_item_headers()

        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            if self.config.config['formatting']['line_items_section']['header_bold']:
                cell.font = Font(bold=True)
            bg_color = self.config.config['formatting']['line_items_section']['header_background_color']
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
        # Add line items
        data_start_row = self.config.line_items_start_row
        for idx, item in enumerate(line_items):
            row = data_start_row + idx

            ws.cell(row=row, column=self.config.get_line_item_column('description'),
                   value=item.get('description', ''))
            ws.cell(row=row, column=self.config.get_line_item_column('amount_excl_vat'),
                   value=item.get('amount_excl_vat', 0))
            ws.cell(row=row, column=self.config.get_line_item_column('vat'),
                   value=item.get('vat', 0))

            # VAT percentage formula using config
            amount_col = get_column_letter(self.config.get_line_item_column('amount_excl_vat'))
            vat_col = get_column_letter(self.config.get_line_item_column('vat'))
            ws.cell(row=row, column=self.config.get_line_item_column('vat_percent'),
                   value=f'=IF({amount_col}{row}=0,0,{vat_col}{row}/{amount_col}{row}*100)')

            ws.cell(row=row, column=self.config.get_line_item_column('total'),
                   value=item.get('total', 0))

            # Deductible checkbox
            deductible_cell = ws.cell(row=row, column=self.config.get_line_item_column('deductible'),
                                    value=item.get('deductible', True))

            # Add checkbox-style data validation
            boolean_options = self.config.get_boolean_validation_options()
            show_dropdown = self.config.get_validation_show_dropdown()
            dv = DataValidation(type="list", formula1=f'"{boolean_options}"', showDropDown=show_dropdown)
            dv.add(deductible_cell)
            ws.add_data_validation(dv)

            # Notes column - add note for non-deductible items
            if not item.get('deductible', True):
                note_text = self.config.get_text_message('non_deductible_note')
                notes_cell = ws.cell(row=row, column=self.config.get_line_item_column('notes'),
                                   value=note_text)
                notes_cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                ws.cell(row=row, column=self.config.get_line_item_column('notes'), value='')
            
    def _add_receipt_image(self, ws: Worksheet, receipt: Dict[str, Any], images_dir: Path):
        """Add receipt image to worksheet"""
        try:
            # Find the image file
            original_file = receipt.get('receipt_info', {}).get('original_file', '')
            if not original_file:
                return
                
            image_path = images_dir / Path(original_file).with_suffix('.jpg').name
            
            if image_path.exists():
                img = XLImage(str(image_path))
                
                # Scale image to fit in merged cells
                img.width, img.height = self.config.get_image_dimensions()
                
                # Position image using dynamic config
                position_cell = self.config.get_image_position_cell()
                ws.add_image(img, position_cell)

                # Merge cells for image area using dynamic config
                merge_range = self.config.get_image_merge_range()
                ws.merge_cells(merge_range)
                
                logger.info(f"Added image to worksheet: {image_path.name}")
        except Exception as e:
            logger.error(f"Error adding image to worksheet: {e}")
            
    def _add_validation_and_formatting(self, ws: Worksheet, receipt: Dict[str, Any]):
        """Add conditional formatting and validation"""

        # Add conditional formatting for VAT validation using config
        vat_range = self.config.get_conditional_formatting_range('vat_validation')
        vat_formula = self.config.get_conditional_formatting_formula('vat_validation')
        vat_yellow_fill = PatternFill(start_color=self.config.get_color('vat_validation'),
                                     end_color=self.config.get_color('vat_validation'),
                                     fill_type="solid")
        vat_rule = FormulaRule(formula=[vat_formula], fill=vat_yellow_fill)
        ws.conditional_formatting.add(vat_range, vat_rule)

        # Red fill for non-deductible items using config
        non_deductible_range = self.config.get_conditional_formatting_range('non_deductible')
        non_deductible_formula = self.config.get_conditional_formatting_formula('non_deductible')
        non_deductible_fill = PatternFill(start_color=self.config.get_color('non_deductible'),
                                        end_color=self.config.get_color('non_deductible'),
                                        fill_type="solid")
        non_deductible_rule = FormulaRule(formula=[non_deductible_formula], fill=non_deductible_fill)
        ws.conditional_formatting.add(non_deductible_range, non_deductible_rule)

        # Add notes for validation errors using config for cell references
        amounts = receipt.get('amounts', {})
        total_excl = amounts.get('total_excl_vat', 0)
        vat = amounts.get('vat_amount', 0)
        total_incl = amounts.get('total_incl_vat', 0)

        # Get cell references for validation checks
        total_incl_cell = self.config.get_header_cell_reference('total_incl_vat')
        verify_cell = self.config.get_cell_reference(
            self.config.header_start_row + 8,  # total_incl_vat row
            self.config.header_value_column + 1  # verification column
        )

        # Check total validation
        if abs((total_excl + vat) - total_incl) > 0.01:
            error_msg = self.config.get_text_message('total_mismatch_error')
            ws[verify_cell] = error_msg
            ws[verify_cell].font = Font(color=self.config.get_color('error'))

        # Check VAT percentage and add warning to VAT amount cell
        if total_excl > 0:
            vat_pct = (vat / total_excl) * 100
            if abs(vat_pct) > 0.1 and abs(vat_pct - 18) > 0.1:
                vat_note_cell = self.config.get_cell_reference(
                    self.config.header_start_row + 6,  # vat_amount row
                    self.config.header_value_column + 2  # notes column
                )
                warning_msg = self.config.get_text_message('vat_warning_format', vat_pct=vat_pct)
                ws[vat_note_cell] = warning_msg
                ws[vat_note_cell].font = Font(color=self.config.get_color('warning'))
                
        
    def _load_categories(self) -> List[str]:
        """Load categories from ICOUNT_CATEGORIES.md file"""
        categories = []
        
        try:
            with open(self.categories_file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            # Extract category names from first column of tables (more restrictive pattern)
            # Matches: | **category_name** | at the start of a line
            pattern = r'^\|\s*\*\*([^*]+)\*\*\s*\|'
            matches = re.findall(pattern, content, re.MULTILINE)
            
            # Filter out table headers
            skip_items = self.config.get_category_skip_items()
            categories = [m.strip() for m in matches if m.strip() not in skip_items]
            
        except Exception as e:
            logger.error(f"Error loading categories from {self.categories_file_path}: {e}")
            
        return categories