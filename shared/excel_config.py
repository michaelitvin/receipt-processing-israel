"""Excel layout configuration loader and utilities"""

import yaml
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
import logging

logger = logging.getLogger(__name__)


class ExcelLayoutConfig:
    """Loads and provides access to Excel layout configuration"""

    def __init__(self, config_path: Optional[Path] = None):
        """Initialize configuration loader

        Args:
            config_path: Path to YAML config file. If None, uses default location.
        """
        if config_path is None:
            # Default config path relative to this file
            config_path = Path(__file__).parent.parent / 'config' / 'excel_layout.yaml'

        self.config_path = config_path
        self.config = self._load_config()

    def _load_config(self) -> Dict[str, Any]:
        """Load configuration from YAML file"""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
            logger.debug(f"Loaded Excel configuration from {self.config_path}")
            return config
        except Exception as e:
            logger.error(f"Failed to load Excel configuration from {self.config_path}: {e}")
            raise

    # Header section properties
    @property
    def header_start_row(self) -> int:
        """First row for header fields"""
        return self.config['header_section']['start_row']

    @property
    def header_field_column(self) -> int:
        """Column for header field names (Hebrew)"""
        return self.config['header_section']['field_column']

    @property
    def header_value_column(self) -> int:
        """Column for header field values"""
        return self.config['header_section']['value_column']

    @property
    def header_max_rows(self) -> int:
        """Maximum rows to search for header fields"""
        return self.config['header_section']['max_rows']

    # Line items section properties
    @property
    def line_items_header_row(self) -> int:
        """Row containing line item column headers"""
        return self.config['line_items_section']['header_row']

    @property
    def line_items_start_row(self) -> int:
        """First row for line item data"""
        return self.config['line_items_section']['data_start_row']

    @property
    def line_items_max_row(self) -> int:
        """Maximum row for line item processing"""
        return self.config['line_items_section']['max_rows']

    def get_line_item_column(self, field: str) -> int:
        """Get column number for a line item field"""
        return self.config['line_items_section']['columns'][field]

    # Field mappings
    def get_field_mappings(self) -> Dict[str, str]:
        """Get Hebrew to English field mappings as dict"""
        mappings = {}
        for mapping in self.config['field_mappings']:
            mappings[mapping['hebrew']] = mapping['english']
        return mappings

    def get_header_fields(self) -> List[Tuple[str, str]]:
        """Get header fields as list of (hebrew, english) tuples"""
        return [(mapping['hebrew'], mapping['english'])
                for mapping in self.config['field_mappings']]

    def get_line_item_headers(self) -> List[str]:
        """Get line item column headers in Hebrew"""
        return self.config['line_item_headers']

    # Formula utilities
    def get_sumif_range(self, range_name: str, start_row: Optional[int] = None,
                       end_row: Optional[int] = None) -> str:
        """Get SUMIF formula range with optional custom start/end rows"""
        if start_row is None:
            start_row = self.line_items_start_row
        if end_row is None:
            end_row = self.line_items_max_row

        # Get field name from new configuration structure
        range_config = self.config['formulas']['sumif_ranges'][range_name]
        field_name = range_config['field']

        return self.get_dynamic_range(field_name, start_row, end_row)

    def get_verification_formula(self, formula_name: str, **kwargs) -> str:
        """Get verification formula with substituted parameters"""
        formula_config = self.config['formulas']['verification'][formula_name]

        if formula_name == 'total_check':
            # For total_check, we need to generate cell references for header fields
            excl_vat_field = formula_config['excl_vat_field']
            vat_field = formula_config['vat_field']

            # Get cell references for these header fields
            excl_vat_cell = self.get_header_cell_reference(excl_vat_field)
            vat_cell = self.get_header_cell_reference(vat_field)

            return f"={excl_vat_cell}+{vat_cell}"

        # For other formula types, use the provided kwargs
        return str(formula_config).format(**kwargs)

    # Conditional formatting utilities
    def get_conditional_formatting_range(self, rule_name: str,
                                       start_row: Optional[int] = None) -> str:
        """Get conditional formatting range"""
        if start_row is None:
            start_row = self.line_items_start_row

        rule_config = self.config['conditional_formatting'][rule_name]
        field_name = rule_config['field']

        if rule_name == 'vat_validation':
            end_row = start_row + rule_config['rows_count'] - 1
            return self.get_dynamic_range(field_name, start_row, end_row)
        elif rule_name == 'non_deductible':
            return self.get_dynamic_range(field_name, start_row, self.line_items_max_row)
        else:
            return self.get_dynamic_range(field_name, start_row, start_row)

    def get_conditional_formatting_formula(self, rule_name: str,
                                         start_row: Optional[int] = None) -> str:
        """Get conditional formatting formula"""
        if start_row is None:
            start_row = self.line_items_start_row

        rule_config = self.config['conditional_formatting'][rule_name]
        field_name = rule_config['field']
        formula_template = rule_config['formula_template']

        # Generate cell reference for the field at start row
        cell_ref = self.get_dynamic_cell_reference(field_name, start_row)

        return formula_template.format(cell=cell_ref)

    # Color utilities
    def get_color(self, color_name: str) -> str:
        """Get color code for formatting"""
        return self.config['colors'][color_name]

    # Cell reference utilities
    def get_cell_reference(self, row: int, column: int) -> str:
        """Convert row/column numbers to Excel cell reference (e.g., A1)"""
        from openpyxl.utils import get_column_letter
        return f"{get_column_letter(column)}{row}"

    def get_column_letter_for_field(self, field_name: str) -> str:
        """Convert field name to Excel column letter using line items column mapping"""
        from openpyxl.utils import get_column_letter
        column_number = self.get_line_item_column(field_name)
        return get_column_letter(column_number)

    def get_dynamic_range(self, field_name: str, start_row: Optional[int] = None,
                         end_row: Optional[int] = None) -> str:
        """Generate Excel range for a field (e.g., 'F21:F121')"""
        if start_row is None:
            start_row = self.line_items_start_row
        if end_row is None:
            end_row = self.line_items_max_row

        column_letter = self.get_column_letter_for_field(field_name)
        return f"{column_letter}{start_row}:{column_letter}{end_row}"

    def get_dynamic_cell_reference(self, field_name: str, row: int) -> str:
        """Generate single cell reference for a field at specific row (e.g., 'D21')"""
        column_letter = self.get_column_letter_for_field(field_name)
        return f"{column_letter}{row}"

    def get_title_cells(self) -> List[str]:
        """Generate title cell references dynamically"""
        title_row = self.config['formatting']['header_section']['title_row']
        title_columns = self.config['formatting']['header_section']['title_columns']

        from openpyxl.utils import get_column_letter
        return [f"{get_column_letter(col)}{title_row}" for col in range(1, title_columns + 1)]

    def get_image_position_cell(self) -> str:
        """Generate image position cell reference dynamically"""
        start_column = self.config['formatting']['image_section']['start_column']
        start_row = self.config['formatting']['image_section']['start_row']

        from openpyxl.utils import get_column_letter
        return f"{get_column_letter(start_column)}{start_row}"

    def get_image_merge_range(self) -> str:
        """Generate image merge range dynamically"""
        start_column = self.config['formatting']['image_section']['start_column']
        start_row = self.config['formatting']['image_section']['start_row']
        merge_columns = self.config['formatting']['image_section']['merge_columns']
        merge_rows = self.config['formatting']['image_section']['merge_rows']

        from openpyxl.utils import get_column_letter
        start_cell = f"{get_column_letter(start_column)}{start_row}"
        end_column = start_column + merge_columns - 1
        end_row = start_row + merge_rows - 1
        end_cell = f"{get_column_letter(end_column)}{end_row}"

        return f"{start_cell}:{end_cell}"

    def get_header_cell_reference(self, field_english: str, value: bool = True) -> str:
        """Get cell reference for a header field

        Args:
            field_english: English field name
            value: If True, return value column. If False, return field name column.
        """
        # Find the row for this field
        for i, (hebrew, english) in enumerate(self.get_header_fields()):
            if english == field_english:
                row = self.header_start_row + i
                column = self.header_value_column if value else self.header_field_column
                return self.get_cell_reference(row, column)

        raise ValueError(f"Field '{field_english}' not found in configuration")

    # Document type methods
    def get_document_types(self) -> List[str]:
        """Get list of valid document types"""
        return self.config['document_types']['valid_types']

    def get_document_type_mapping(self, english_type: str) -> str:
        """Map English document type to Hebrew"""
        mappings = self.config['document_types']['mappings']
        return mappings.get(english_type, english_type)

    # Column width methods
    def get_column_widths(self) -> Dict[str, Any]:
        """Get all column width configurations"""
        return self.config['column_widths']

    def get_header_column_width(self, column_type: str) -> int:
        """Get width for a specific header column type"""
        return self.config['column_widths']['header_section'].get(column_type, 15)

    # Header title methods
    def get_header_titles(self) -> List[str]:
        """Get header title labels"""
        return self.config['header_titles']

    # Image settings methods
    def get_image_dimensions(self) -> Tuple[int, int]:
        """Get image width and height"""
        settings = self.config['image_settings']
        return settings['width'], settings['height']

    # Worksheet settings methods
    def get_worksheet_name(self, idx: int) -> str:
        """Generate worksheet name for given index"""
        settings = self.config['worksheet_settings']
        prefix = settings['name_prefix']
        name_format = settings['name_format']
        return name_format.format(prefix=prefix, idx=idx)

    def get_default_sheet_name(self) -> str:
        """Get default Excel sheet name"""
        return self.config['worksheet_settings']['default_sheet_name']

    # Validation settings methods
    def get_boolean_validation_options(self) -> str:
        """Get boolean validation options string"""
        return self.config['validation_settings']['boolean_options']

    def get_validation_show_dropdown(self) -> bool:
        """Get whether to show dropdown for validation"""
        return self.config['validation_settings']['show_dropdown']

    # Text message methods
    def get_text_message(self, message_key: str, **kwargs) -> str:
        """Get text message by key, with optional formatting"""
        message = self.config['text_messages'].get(message_key, '')
        if kwargs and message:
            return message.format(**kwargs)
        return message

    # Category settings methods
    def get_category_skip_items(self) -> List[str]:
        """Get items to skip when loading categories"""
        return self.config['category_settings']['skip_items']


# Global instance for easy access
_config_instance = None

def get_excel_config() -> ExcelLayoutConfig:
    """Get global configuration instance (singleton)"""
    global _config_instance
    if _config_instance is None:
        _config_instance = ExcelLayoutConfig()
    return _config_instance