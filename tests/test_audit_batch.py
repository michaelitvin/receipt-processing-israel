import json
import subprocess
import sys
from pathlib import Path

import pytest
from PIL import Image

REPO = Path(__file__).parent.parent
sys.path.insert(0, str(REPO))
sys.path.insert(0, str(REPO / "tools"))

from shared.excel_generator import ExcelGenerator
from audit_batch import parse_batch


def build_receipt(name, number, date, total, line_items, vendor="ספק בדיקה", vendor_id="510"):
    return {
        "status": "success",
        "receipt_info": {"number": number, "vendor": vendor, "vendor_id": vendor_id,
                         "date": date, "document_type": "invoice+receipt",
                         "original_file": name, "reasoning": "בדיקה", "currency": "ILS"},
        "amounts": {"total_excl_vat": round(total / 1.18, 2),
                    "vat_amount": round(total - total / 1.18, 2),
                    "total_incl_vat": total},
        "classification": {"category": "דלק", "confidence": 0.9},
        "line_items": line_items,
    }


@pytest.fixture(scope="module")
def batch(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("batch")
    images = tmp / "images"
    images.mkdir()
    for stem in ("one", "two"):
        Image.new("RGB", (800, 1200), "white").save(images / f"{stem}.jpg")
    receipts = [
        build_receipt("one.pdf", "111", "2026-05-10", 118.0,
                      [{"description": "פריט", "amount_excl_vat": 100.0, "vat": 18.0,
                        "total": 118.0, "deductible": True}]),
        build_receipt("two.pdf", "", "2022-01-01", 0,
                      [{"description": "ריק", "amount_excl_vat": 0, "vat": 0,
                        "total": 0, "deductible": True}]),
    ]
    gen = ExcelGenerator(REPO / "docs" / "extraction-prompt" / "001-ICOUNT_CATEGORIES.md")
    wb = gen.create_batch_workbook(receipts, images)
    xlsx = tmp / "batch.xlsx"
    wb.save(xlsx)
    return xlsx


def run_cli(*args):
    result = subprocess.run(
        ["uv", "run", "python", str(REPO / "tools" / "audit_batch.py"), *map(str, args)],
        capture_output=True, cwd=REPO)
    assert result.returncode in (0, 1), result.stderr.decode("utf-8", "replace")
    return json.loads(result.stdout.decode("utf-8")), result.returncode


def test_parse_batch_round_trips_values(batch):
    receipts = parse_batch(batch)
    assert len(receipts) == 2
    r1 = receipts[0]
    assert r1["sheet"] == "R001"
    assert r1["receipt_info"]["number"] == "111"
    assert r1["receipt_info"]["vendor"] == "ספק בדיקה"
    assert float(r1["amounts"]["total_incl_vat"]) == 118.0
    assert r1["line_items"][0]["description"] == "פריט"
    assert r1["line_items"][0]["deductible"] is True
    assert r1["image_jpg"].endswith("one.jpg")
    # the generator's sum row must NOT be parsed as a line item
    assert all(li["description"] != 'סה"כ פריטים' for li in r1["line_items"])


def test_manifest_cli(batch):
    manifest, _ = run_cli("manifest", batch)
    assert manifest["R001"]["receipt_info"]["number"] == "111"
    assert manifest["R002"]["amounts"]["total_incl_vat"] in (0, 0.0, "0")


def test_check_cli_flags_bad_sheet_only(batch):
    issues, rc = run_cli("check", batch, "--period", "2026-05")
    assert "R001" not in issues
    r2 = issues["R002"]
    assert any("חסר מספר קבלה" in w for w in r2)
    assert any("מחוץ לתקופת הדיווח" in w for w in r2)
    assert rc == 1  # issues found -> exit 1


def test_renamed_sheets_still_parse(batch, tmp_path):
    # A reviewed batch may carry vendor suffixes on sheet names (R001partner).
    from openpyxl import load_workbook
    work = tmp_path / "renamed.xlsx"
    import shutil as sh
    sh.copy2(batch, work)
    wb = load_workbook(work)
    wb["R001"].title = "R001partner"
    wb["R002"].title = "R002gas"
    wb.save(work)
    receipts = parse_batch(work)
    assert {r["sheet"] for r in receipts} == {"R001partner", "R002gas"}


def test_check_fails_loudly_when_no_receipt_sheets(batch, tmp_path):
    # Sheets that don't match R### must not be silently skipped to a green result.
    from openpyxl import load_workbook
    work = tmp_path / "nosheets.xlsx"
    import shutil as sh
    sh.copy2(batch, work)
    wb = load_workbook(work)
    wb["R001"].title = "junk1"
    wb["R002"].title = "junk2"
    wb.save(work)
    result = subprocess.run(
        ["uv", "run", "python", str(REPO / "tools" / "audit_batch.py"), "check", str(work)],
        capture_output=True, cwd=REPO)
    assert result.returncode == 2, result.stderr.decode("utf-8", "replace")
    payload = json.loads(result.stdout.decode("utf-8"))
    assert "workbook" in payload


def test_agent_prompts_values_from_manifest(batch):
    prompts, _ = run_cli("agent-prompts", batch, "--chunk", "1")
    assert len(prompts) == 2  # 2 receipts, chunk size 1
    p1 = prompts[0]["prompt"]
    assert "111" in p1 and "total=118" in p1 and "ספק בדיקה" in p1
    assert "one.jpg" in p1
    # anti-anchoring: transcription instruction precedes the extracted values
    assert p1.index("TRANSCRIBE") < p1.index("extracted values for comparison")
    assert prompts[0]["label"] == "audit:R001"


def test_agent_prompts_chunking(batch):
    prompts, _ = run_cli("agent-prompts", batch, "--chunk", "6")
    assert len(prompts) == 1
    assert prompts[0]["label"] == "audit:R001-R002"


def test_apply_fixes_and_verify(batch, tmp_path):
    import shutil as sh
    work = tmp_path / "work.xlsx"
    sh.copy2(batch, work)
    fixes = [
        {"sheet": "R001", "field": "vendor", "value": "ספק מתוקן", "note": "תוקן בביקורת"},
        {"sheet": "R002", "line_item": 0,
         "values": {"description": "דלק", "amount_excl_vat": 100.0, "vat": 18.0, "total": 118.0}},
        {"sheet": "R002", "field": "total_incl_vat", "value": 118.0},
        {"sheet": "R002", "field": "total_excl_vat", "value": 100.0},
        {"sheet": "R002", "field": "vat_amount", "value": 18.0},
        {"sheet": "R002", "non_expense": True, "note": "לא הוצאה"},
    ]
    fixes_path = tmp_path / "fixes.json"
    fixes_path.write_text(json.dumps(fixes, ensure_ascii=False), encoding="utf-8")
    backup_dir = tmp_path / "backups"

    out, rc = run_cli("apply-fixes", work, fixes_path, "--backup-dir", backup_dir)
    assert rc == 0
    assert out["applied"] == 6
    assert list(backup_dir.glob("*.xlsx")), "backup must exist"

    from openpyxl import load_workbook
    wb = load_workbook(work)
    from shared.excel_config import get_excel_config
    config = get_excel_config()
    fields = [f for _, f in config.get_header_fields()]
    vendor_row = config.header_start_row + fields.index("vendor")
    ws1 = wb["R001"]
    assert ws1.cell(row=vendor_row, column=config.header_value_column).value == "ספק מתוקן"
    note = ws1.cell(row=vendor_row, column=config.header_value_column + 2).value
    assert note == "תוקן בביקורת"
    ws2 = wb["R002"]
    assert "FF0000" in str(ws2.sheet_properties.tabColor.rgb)
    li_row = config.line_items_start_row
    assert ws2.cell(row=li_row, column=config.get_line_item_column("description")).value == "דלק"

    report, rc = run_cli("verify", work)
    assert rc == 0, report


def test_verify_catches_broken_arithmetic(batch, tmp_path):
    import shutil as sh
    work = tmp_path / "broken.xlsx"
    sh.copy2(batch, work)
    fixes = [{"sheet": "R001", "field": "total_incl_vat", "value": 999.0}]
    fixes_path = tmp_path / "fixes.json"
    fixes_path.write_text(json.dumps(fixes), encoding="utf-8")
    run_cli("apply-fixes", work, fixes_path, "--backup-dir", tmp_path / "b")
    report, rc = run_cli("verify", work)
    assert rc == 1
    assert any("R001" in k for k in report)
