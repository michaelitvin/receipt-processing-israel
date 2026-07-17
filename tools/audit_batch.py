#!/usr/bin/env python3
# Copyright (c) 2025 Michael Litvin
# Licensed under AGPL-3.0-or-later - see LICENSE file for details
"""Deterministic audit tooling for extraction batch workbooks.

Subcommands (JSON in/out, UTF-8):
  manifest <xlsx>                       dump per-sheet receipt data
  check <xlsx> [--period YYYY-MM]      structural checks (exit 1 if issues)
  agent-prompts <xlsx> [--chunk N]     visual-verification agent prompts
  apply-fixes <xlsx> <fixes.json> --backup-dir DIR
  verify <xlsx>                        post-fix integrity (exit 1 if broken)

fixes.json is a JSON list; each entry is one of:
  {"sheet": "R001", "field": "vendor", "value": "X", "note": "optional"}
  {"sheet": "R011", "line_item": 0,
   "values": {"description": "...", "amount_excl_vat": 1.0, "vat": 0.2, "total": 1.2},
   "note": "optional"}
  {"sheet": "R022", "non_expense": true, "note": "required"}

Used by the bimonthly-cycle skill; see .claude/skills/bimonthly-cycle/SKILL.md.
"""
import argparse
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from shared.excel_config import get_excel_config
from shared.receipt_checks import check_batch, parse_period

# Sheet names may carry a human-added vendor suffix after review (e.g. "R001partner")
SHEET_RE = re.compile(r"R\d{3}.*")
LINE_ITEMS_SUM_LABEL = 'סה"כ פריטים'
AUDIT_NOTE_COLOR = "CC5500"

INFO_FIELDS = {"number", "vendor", "vendor_id", "date", "document_type",
               "currency", "original_file", "reasoning"}
AMOUNT_FIELDS = {"total_excl_vat", "vat_amount", "total_incl_vat"}


def parse_batch(xlsx: Path) -> list:
    """Parse an extraction batch workbook back into standard receipt dicts.

    Each dict gains extra keys: sheet, source_pdf, image_jpg.
    """
    config = get_excel_config()
    wb = load_workbook(xlsx)
    value_col = config.header_value_column
    receipts = []

    for name in wb.sheetnames:
        if not SHEET_RE.fullmatch(name):
            continue
        ws = wb[name]
        info, amounts, classification = {}, {}, {}
        source_pdf = ""

        row = config.header_start_row
        for _, field in config.get_header_fields():
            cell = ws.cell(row=row, column=value_col)
            value = cell.value
            if field in INFO_FIELDS:
                info[field] = "" if value is None else value
            elif field in AMOUNT_FIELDS:
                amounts[field] = 0 if value is None else value
            elif field == "category":
                classification["category"] = "" if value is None else value
            if field == "original_file" and cell.hyperlink:
                source_pdf = re.sub(r"^file:///", "", cell.hyperlink.target).replace("/", "\\")
            row += 1

        line_items = []
        li_row = config.line_items_start_row
        while li_row <= ws.max_row:
            desc = ws.cell(row=li_row, column=config.get_line_item_column("description")).value
            if desc is None or not str(desc).strip() or str(desc).strip() == LINE_ITEMS_SUM_LABEL:
                break
            line_items.append({
                "description": str(desc).strip(),
                "amount_excl_vat": ws.cell(row=li_row, column=config.get_line_item_column("amount_excl_vat")).value or 0,
                "vat": ws.cell(row=li_row, column=config.get_line_item_column("vat")).value or 0,
                "total": ws.cell(row=li_row, column=config.get_line_item_column("total")).value or 0,
                "deductible": bool(ws.cell(row=li_row, column=config.get_line_item_column("deductible")).value),
            })
            li_row += 1

        image_jpg = ""
        if info.get("original_file"):
            image_jpg = str(xlsx.parent / "images" / (Path(str(info["original_file"])).stem + ".jpg"))

        receipts.append({
            "sheet": name,
            "receipt_info": info,
            "amounts": amounts,
            "classification": classification,
            "line_items": line_items,
            "source_pdf": source_pdf,
            "image_jpg": image_jpg,
        })
    return receipts


def _require_sheets(receipts: list) -> None:
    """Guard against a workbook whose receipt sheets none of us can see.

    A reviewed batch with renamed/reordered sheets used to parse to zero receipts
    and pass every check silently. Fail loudly instead.
    """
    if not receipts:
        print(json.dumps(
            {"workbook": ["no R### receipt sheets parsed - wrong file or unreadable layout"]},
            ensure_ascii=False, indent=1))
        raise SystemExit(2)


def cmd_manifest(args) -> int:
    receipts = parse_batch(args.xlsx)
    _require_sheets(receipts)
    print(json.dumps({r["sheet"]: r for r in receipts}, ensure_ascii=False, indent=1, default=str))
    return 0


def cmd_check(args) -> int:
    receipts = parse_batch(args.xlsx)
    _require_sheets(receipts)
    period_months = parse_period(args.period) if args.period else None
    warnings = check_batch(receipts, period_months)
    issues = {receipts[i]["sheet"]: w for i, w in warnings.items() if w}
    print(json.dumps(issues, ensure_ascii=False, indent=1))
    return 1 if issues else 0


PROMPT_HEADER = """You are auditing extracted receipt data against the actual receipt \
images for an Israeli tax-reporting pipeline. Hebrew receipts read right-to-left.

For each receipt below, follow this exact order:
1. Read the JPG image with the Read tool and TRANSCRIBE the key printed values \
BEFORE looking at the extracted values: vendor name, vendor tax id (ח.פ/עוסק), \
receipt/invoice number, date, net amount, VAT, total, currency, document type \
(חשבונית / קבלה / חשבונית מס קבלה). Report your transcription.
2. Only then compare your transcription against the extracted values listed for \
that receipt and note real differences.
3. Judge whether the page-1 JPG is representative for human review: does it show \
the vendor and total, or is it a cover page/email with the invoice on a later page?

If the page-1 JPG lacks the needed values, render more pages of the source PDF \
(run from {repo}):
  uv run python -c "import pdf2image; [im.save(rf'{scratch}\\audit_page_{{i}}.jpg','JPEG') \
for i,im in enumerate(pdf2image.convert_from_path(r'<SOURCE_PDF>', dpi=150))]"
then Read those page JPGs.

Do NOT modify any files except temp page renders in the scratch directory.

"""

PROMPT_FOOTER = """
Return one block per receipt, exactly this shape:
SHEET_NAME: VERDICT (OK / MISMATCH / IMAGE-NOT-REPRESENTATIVE / UNVERIFIABLE)
- transcribed: <the values you read off the image>
- mismatches: <field: extracted vs printed - only real differences, or 'none'>
- image_representative: yes/no + what page 1 shows
- notes: <anything odd: possible duplicate, wrong doc type, non-expense document, \
unusual billing entity, unreadable areas>"""


def _receipt_prompt_block(r: dict) -> str:
    info, amounts = r["receipt_info"], r["amounts"]
    return (
        f"{r['sheet']}:\n"
        f"  image: {r['image_jpg']}\n"
        f"  source_pdf: {r['source_pdf']}\n"
        f"  extracted values for comparison (step 2): "
        f"receipt_number={info.get('number')!r}, vendor={info.get('vendor')!r}, "
        f"vendor_id={info.get('vendor_id')!r}, date={info.get('date')!r}, "
        f"doc_type={info.get('document_type')!r}, currency={info.get('currency')!r}, "
        f"net={amounts.get('total_excl_vat')}, vat={amounts.get('vat_amount')}, "
        f"total={amounts.get('total_incl_vat')}, "
        f"category={r['classification'].get('category')!r}\n"
        f"  extracted line items: "
        + "; ".join(
            f"{li['description']} (net={li['amount_excl_vat']}, vat={li['vat']}, "
            f"total={li['total']}, deductible={li['deductible']})"
            for li in r["line_items"]) + "\n"
    )


def cmd_agent_prompts(args) -> int:
    receipts = parse_batch(args.xlsx)
    repo = Path(__file__).parent.parent
    header = PROMPT_HEADER.format(repo=repo, scratch=args.scratch)
    prompts = []
    for i in range(0, len(receipts), args.chunk):
        chunk = receipts[i:i + args.chunk]
        label = (f"audit:{chunk[0]['sheet']}" if len(chunk) == 1
                 else f"audit:{chunk[0]['sheet']}-{chunk[-1]['sheet']}")
        body = "\n".join(_receipt_prompt_block(r) for r in chunk)
        prompts.append({"label": label, "prompt": header + body + PROMPT_FOOTER})
    print(json.dumps(prompts, ensure_ascii=False, indent=1))
    return 0


def cmd_apply_fixes(args) -> int:
    fixes = json.loads(Path(args.fixes).read_text(encoding="utf-8"))
    config = get_excel_config()
    field_rows = {field: config.header_start_row + i
                  for i, (_, field) in enumerate(config.get_header_fields())}
    value_col = config.header_value_column
    notes_col = value_col + 2
    note_font = Font(color=AUDIT_NOTE_COLOR, bold=True)

    args.backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = args.backup_dir / f"{args.xlsx.stem}.backup-{stamp}.xlsx"
    shutil.copy2(args.xlsx, backup)

    wb = load_workbook(args.xlsx)
    applied = 0
    for fix in fixes:
        ws = wb[fix["sheet"]]
        note_row = None
        if fix.get("non_expense"):
            ws.sheet_properties.tabColor = "FF0000"
            note_row = config.header_start_row
        elif "field" in fix:
            note_row = field_rows[fix["field"]]
            ws.cell(row=note_row, column=value_col, value=fix["value"])
        elif "line_item" in fix:
            row = config.line_items_start_row + fix["line_item"]
            for key, value in fix["values"].items():
                ws.cell(row=row, column=config.get_line_item_column(key), value=value)
        else:
            raise ValueError(f"Unrecognized fix entry: {fix}")
        if fix.get("note") and note_row is not None:
            cell = ws.cell(row=note_row, column=notes_col, value=fix["note"])
            cell.font = note_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        applied += 1

    try:
        wb.save(args.xlsx)
    except PermissionError:
        print(f"Cannot save {args.xlsx} - the file is open in Excel. "
              f"Close it and re-run. (Backup kept at {backup})", file=sys.stderr)
        return 3
    print(json.dumps({"applied": applied, "backup": str(backup)}, ensure_ascii=False))
    return 0


def cmd_verify(args) -> int:
    from shared.receipt_checks import AMOUNT_TOLERANCE, _num
    config = get_excel_config()
    wb = load_workbook(args.xlsx)
    problems = {}

    if "CategoryList" not in wb.defined_names:
        problems["workbook"] = ["missing CategoryList named range"]

    fields = [f for _, f in config.get_header_fields()]
    value_col = config.header_value_column
    net_row = config.header_start_row + fields.index("total_excl_vat")
    vat_row = config.header_start_row + fields.index("vat_amount")
    total_row = config.header_start_row + fields.index("total_incl_vat")

    checked = 0
    for name in wb.sheetnames:
        if not SHEET_RE.fullmatch(name):
            continue
        checked += 1
        ws = wb[name]
        sheet_problems = []
        if len(getattr(ws, "_images", [])) != 1:
            sheet_problems.append(f"{len(ws._images)} embedded images (expected 1)")
        if not ws.data_validations.dataValidation:
            sheet_problems.append("no data validations")
        if not any(c.hyperlink for row in ws.iter_rows() for c in row):
            sheet_problems.append("no source hyperlink")
        net = _num(ws.cell(row=net_row, column=value_col).value)
        vat = _num(ws.cell(row=vat_row, column=value_col).value)
        total = _num(ws.cell(row=total_row, column=value_col).value)
        if total and abs(net + vat - total) > AMOUNT_TOLERANCE:
            sheet_problems.append(f"arithmetic: {net:g} + {vat:g} != {total:g}")
        if sheet_problems:
            problems[name] = sheet_problems

    if not checked:
        problems["workbook"] = ["no R### receipt sheets found - wrong file or unreadable layout"]

    print(json.dumps(problems, ensure_ascii=False, indent=1))
    return 1 if problems else 0


def main() -> int:
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8")
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("manifest", help="dump per-sheet receipt data as JSON")
    p.add_argument("xlsx", type=Path)
    p.set_defaults(func=cmd_manifest)

    p = sub.add_parser("check", help="run structural checks")
    p.add_argument("xlsx", type=Path)
    p.add_argument("--period", help="reporting period YYYY-MM")
    p.set_defaults(func=cmd_check)

    p = sub.add_parser("agent-prompts", help="emit visual-verification agent prompts")
    p.add_argument("xlsx", type=Path)
    p.add_argument("--chunk", type=int, default=6)
    p.add_argument("--scratch", default=str(Path.home() / "AppData" / "Local" / "Temp"),
                   help="directory agents may write temp page renders to")
    p.set_defaults(func=cmd_agent_prompts)

    p = sub.add_parser("apply-fixes", help="apply a fixes.json to the workbook")
    p.add_argument("xlsx", type=Path)
    p.add_argument("fixes", type=Path)
    p.add_argument("--backup-dir", type=Path, required=True)
    p.set_defaults(func=cmd_apply_fixes)

    p = sub.add_parser("verify", help="post-fix integrity check")
    p.add_argument("xlsx", type=Path)
    p.set_defaults(func=cmd_verify)

    args = parser.parse_args()
    if not args.xlsx.exists():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 2
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
